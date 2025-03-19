from flask import Flask, render_template, request, send_from_directory, redirect, url_for, send_file
from openpyxl import load_workbook
from io import BytesIO
import os, xlrd


app = Flask(__name__)

def xlsx_to_txt(file):
 file_extension = os.path.splitext(file.filename)[1].lower()
 output = BytesIO()
 if file_extension == '.xlsx':
  workbook = load_workbook(file, data_only=True)
  for sheet_name in workbook.sheetnames:
   sheet = workbook[sheet_name]
   for row in sheet.iter_rows(values_only=True):
    row_filtered = [str(cell) for cell in row if cell is not None]
    row_str = '\t'.join(row_filtered)
    output.write(row_str.encode('utf-8') + b'\n')
 elif file_extension == '.xls':
  workbook = xlrd.open_workbook(file_contents=file.read())
  for sheet_name in workbook.sheet_names():
   sheet = workbook.sheet_by_name(sheet_name)
   for row_idx in range(sheet.nrows):
    row = sheet.row_values(row_idx)
    row_filtered = [str(cell) for cell in row if cell is not None]
    row_str = '\t'.join(row_filtered)
    output.write(row_str.encode('utf-8') + b'\n')
 else:
  raise ValueError("Formato de archivo no admitido. Use .xlsx o .xls.")
 output.seek(0)
 return output

@app.route('/', methods=['GET', 'POST'])
def index():
 return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
 if 'files[]' not in request.files:
  return "No se seleccionaron archivos", 400
 files = request.files.getlist('files[]')
 for file in files:
  if file.filename == '':
   continue
  try:
   txt_output = xlsx_to_txt(file)
   return send_file(
    txt_output,
    as_attachment=True,
    download_name=os.path.splitext(file.filename)[0] + '.txt',
    mimetype='text/plain'
   )
  except Exception as e:
   return f"Error al convertir {file.filename}: {str(e)}", 500

if __name__ == '__main__':
 app.run(debug=True, port=5000)
